import streamlit as st
import pdfplumber
import re
import zipfile
import os
import pandas as pd
from io import BytesIO
from tempfile import TemporaryDirectory
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Extrator de Notas Fiscais", layout="centered")
st.title("📄 Extrator de Dados de PDFs (Notas Fiscais de Energia)")

st.markdown("Envie múltiplos arquivos `.pdf` ou um `.zip` contendo vários PDFs.")

uploaded_files = st.file_uploader("Upload de PDFs ou ZIP", type=["pdf", "zip"], accept_multiple_files=True)

def extrair_dados_pdf(file):
    with pdfplumber.open(file) as pdf:
        texto = ""
        for page in pdf.pages:
            texto += page.extract_text() + "\n"

    def buscar(regex, flags=0):
        match = re.search(regex, texto, flags)
        return match.group(1).strip() if match else None

    unidade_consumidora = buscar(r'\b(\d{9})\b')

    return {
        "Nota Fiscal": buscar(r'NOTA FISCAL Nº (\d+)'),
        "Série": buscar(r'NOTA FISCAL Nº \d+\s*-\s*SÉRIE\s*(\S+)'),
        "CNPJ": buscar(r'CNPJ/CPF:\s*([\d./-]+)'),
        "Valor (R$)": buscar(r'(\d{1,3},\d{2})\nO Pagamento poderá ser realizado'),
        "Data de Emissão": buscar(r'DATA DE EMISSÃO:\s*(\d{2}/\d{2}/\d{4})'),
        "Nome do Destinatário": buscar(r'^\s*(ROMA HOTEIS.*FILIAL VILLAS)', re.MULTILINE),
        "Protocolo de Autorização": buscar(r'Protocolo de autorização:\s*(.*?)\s*-'),
        "Unidade Consumidora": unidade_consumidora,
        "Chave de Acesso": buscar(r'chave de acesso:\s*([\d]+)')
    }

def processar_arquivos(files):
    dados_extraidos = []

    for file in files:
        if file.name.endswith(".pdf"):
            dados = extrair_dados_pdf(file)
            dados_extraidos.append(dados)

        elif file.name.endswith(".zip"):
            with TemporaryDirectory() as tmpdir:
                zip_path = os.path.join(tmpdir, file.name)
                with open(zip_path, "wb") as f:
                    f.write(file.read())

                with zipfile.ZipFile(zip_path, "r") as zip_ref:
                    zip_ref.extractall(tmpdir)
                    for nome_arquivo in zip_ref.namelist():
                        if nome_arquivo.endswith(".pdf"):
                            caminho_pdf = os.path.join(tmpdir, nome_arquivo)
                            with open(caminho_pdf, "rb") as f_pdf:
                                dados = extrair_dados_pdf(f_pdf)
                                dados_extraidos.append(dados)

    return dados_extraidos

if uploaded_files:
    resultado = processar_arquivos(uploaded_files)
    df_resultado = pd.DataFrame(resultado)

    # Converte valores para float
    df_resultado["Valor (R$)"] = df_resultado["Valor (R$)"].str.replace(",", ".").astype(float)

    st.success("✅ Dados extraídos com sucesso!")
    st.dataframe(df_resultado)

    # Gera planilha formatada
    output = BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Notas Fiscais"

    for r in dataframe_to_rows(df_resultado, index=False, header=True):
        ws.append(r)

    max_col_letter = get_column_letter(ws.max_column)
    table_ref = f"A1:{max_col_letter}{ws.max_row}"
    tab = Table(displayName="NotasFiscais", ref=table_ref)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)

    # Formata a coluna "Chave de Acesso" como texto
    for idx, cell in enumerate(ws[1], 1):
        if cell.value == "Chave de Acesso":
            chave_col_idx = idx
            break

    for row in ws.iter_rows(min_row=2, min_col=chave_col_idx, max_col=chave_col_idx):
        for cell in row:
            cell.number_format = "@"

    wb.save(output)
    st.download_button("📥 Baixar Excel Formatado", data=output.getvalue(), file_name="notas_fiscais_formatado.xlsx")